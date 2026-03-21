"""
Import data from Excel or JSON files into the database.

Supports data from export_data command or manually created files.

Usage:
    python manage.py import_data --file=backup.xlsx
    python manage.py import_data --file=backup.json
    python manage.py import_data --file=data.xlsx --clear  # Clears old data first
"""

import json
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from datetime import datetime
from core.models import Event, Customer, Order, OrderItem, StatusLog, UserProfile


class Command(BaseCommand):
    help = 'Import data from Excel or JSON files'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to Excel or JSON file',
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Clear existing data before import',
        )

    def handle(self, *args, **options):
        file_path = Path(options['file'])
        
        if not file_path.exists():
            raise CommandError(f'File not found: {file_path}')
        
        if options['clear']:
            self.clear_data()
        
        suffix = file_path.suffix.lower()
        
        if suffix == '.xlsx':
            self.import_excel(file_path)
        elif suffix == '.json':
            self.import_json(file_path)
        else:
            raise CommandError(f'Unsupported file format: {suffix}')

    def clear_data(self):
        """Clear existing data (except users)."""
        self.stdout.write(self.style.WARNING('⚠️  Clearing existing data...'))
        with transaction.atomic():
            OrderItem.objects.all().delete()
            StatusLog.objects.all().delete()
            Order.objects.all().delete()
            Customer.objects.all().delete()
            Event.objects.all().delete()
        self.stdout.write(self.style.SUCCESS('✅ Data cleared'))

    def import_excel(self, file_path):
        """Import from Excel file."""
        try:
            wb = load_workbook(file_path)
            
            with transaction.atomic():
                if 'Events' in wb.sheetnames:
                    self.import_events_sheet(wb['Events'])
                
                if 'Customers' in wb.sheetnames:
                    self.import_customers_sheet(wb['Customers'])
                
                if 'Orders' in wb.sheetnames:
                    self.import_orders_sheet(wb['Orders'])
                
                if 'StatusLogs' in wb.sheetnames:
                    self.import_statuslogs_sheet(wb['StatusLogs'])
            
            self.stdout.write(
                self.style.SUCCESS(f'✅ Data imported from {file_path.name}')
            )
        except Exception as e:
            raise CommandError(f'Failed to import from {file_path}: {str(e)}')

    def import_events_sheet(self, ws):
        """Import Events from worksheet."""
        events_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                event, created = Event.objects.get_or_create(
                    id=row[0],
                    defaults={
                        'name': row[1],
                        'date': row[2] if isinstance(row[2], datetime) else datetime.fromisoformat(str(row[2])).date(),
                        'description': row[3] or '',
                        'is_active': row[4] if row[4] is not None else True,
                    }
                )
                if created:
                    events_created += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠️  Event {row[1]}: {str(e)}'))
        
        if events_created:
            self.stdout.write(f'  ✅ {events_created} events imported')

    def import_customers_sheet(self, ws):
        """Import Customers from worksheet."""
        customers_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                customer, created = Customer.objects.get_or_create(
                    id=row[0],
                    defaults={
                        'name': row[1],
                        'phone': row[2] or '',
                        'email': row[3] or '',
                    }
                )
                if created:
                    customers_created += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠️  Customer {row[1]}: {str(e)}'))
        
        if customers_created:
            self.stdout.write(f'  ✅ {customers_created} customers imported')

    def import_orders_sheet(self, ws):
        """Import Orders from worksheet."""
        orders_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                # Get event and customer
                event = Event.objects.get(name=row[1]) if row[1] else None
                customer = Customer.objects.get(name=row[2]) if row[2] else None
                
                if not event or not customer:
                    self.stdout.write(
                        self.style.WARNING(f'  ⚠️  Order #{row[3]}: Event or Customer not found')
                    )
                    continue
                
                order, created = Order.objects.get_or_create(
                    id=row[0],
                    defaults={
                        'event': event,
                        'customer': customer,
                        'ticket_number': row[3],
                        'qr_token': row[4],
                        'forfait': row[5],
                        'nb_persons': row[6],
                        'dining_type': row[7],
                        'meat': row[8],
                        'side': row[9],
                        'vegetable': row[10] or '',
                        'unit_price': row[11],
                        'total_amount': row[12],
                        'payment_status': row[13],
                        'preparation_status': row[14],
                        'started_at': self._parse_datetime(row[15]),
                    }
                )
                if created:
                    orders_created += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠️  Order #{row[3]}: {str(e)}'))
        
        if orders_created:
            self.stdout.write(f'  ✅ {orders_created} orders imported')

    def import_statuslogs_sheet(self, ws):
        """Import StatusLogs from worksheet."""
        logs_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                # Find order by ticket number
                ticket_str = str(row[2]).replace('#', '').strip()
                order = Order.objects.filter(ticket_number=ticket_str).first()
                
                if not order:
                    continue
                
                log, created = StatusLog.objects.get_or_create(
                    id=row[0],
                    defaults={
                        'order': order,
                        'status_type': row[2],
                        'old_status': row[3],
                        'new_status': row[4],
                    }
                )
                if created:
                    logs_created += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠️  StatusLog: {str(e)}'))
        
        if logs_created:
            self.stdout.write(f'  ✅ {logs_created} status logs imported')

    def import_json(self, file_path):
        """Import from JSON file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if 'events' in data:
                with transaction.atomic():
                    self._import_json_events(data['events'])

            if 'customers' in data:
                with transaction.atomic():
                    self._import_json_customers(data['customers'])

            if 'orders' in data:
                with transaction.atomic():
                    self._import_json_orders(data['orders'])

            if 'order_items' in data:
                self._import_json_order_items(data['order_items'])

            if 'statuslogs' in data:
                with transaction.atomic():
                    self._import_json_statuslogs(data['statuslogs'])
            
            self.stdout.write(
                self.style.SUCCESS(f'✅ Data imported from {file_path.name}')
            )
        except Exception as e:
            raise CommandError(f'Failed to import from {file_path}: {str(e)}')

    def _import_json_events(self, events):
        """Import events from JSON data."""
        created = 0
        for item in events:
            try:
                event, is_new = Event.objects.get_or_create(
                    name=item.get('name'),
                    defaults={
                        'date': self._parse_date(item.get('date')),
                        'description': item.get('description', ''),
                        'is_active': item.get('is_active', True),
                    }
                )
                if is_new:
                    created += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠️  Event: {str(e)}'))

        if created:
            self.stdout.write(f'  ✅ {created} events imported')

    def _import_json_customers(self, customers):
        """Import customers from JSON data."""
        created = 0
        for item in customers:
            try:
                customer, is_new = Customer.objects.get_or_create(
                    name=item.get('name'),
                    defaults={
                        'phone': item.get('phone', ''),
                        'email': item.get('email', ''),
                    }
                )
                if is_new:
                    created += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠️  Customer: {str(e)}'))

        if created:
            self.stdout.write(f'  ✅ {created} customers imported')

    def _import_json_orders(self, orders):
        """Import orders from JSON data."""
        created = 0
        for item in orders:
            try:
                event = Event.objects.filter(name=item.get('event')).first()
                customer_name = item.get('customer')
                customer = Customer.objects.filter(name=customer_name).first()

                if not event or not customer:
                    continue

                defaults = {
                    'customer': customer,
                    'forfait': item.get('forfait'),
                    'nb_persons': item.get('nb_persons', 1),
                    'dining_type': item.get('dining_type'),
                    'meat': item.get('meat', ''),
                    'side': item.get('side', ''),
                    'vegetable': item.get('vegetable', ''),
                    'unit_price': item.get('unit_price'),
                    'total_amount': item.get('total_amount', item.get('total_price', 0)),
                    'payment_status': item.get('payment_status'),
                    'preparation_status': item.get('preparation_status'),
                    'started_at': self._parse_datetime(item.get('started_at')),
                }
                if item.get('qr_token'):
                    defaults['qr_token'] = item['qr_token']

                order, is_new = Order.objects.get_or_create(
                    event=event,
                    ticket_number=item.get('ticket_number'),
                    defaults=defaults,
                )
                if is_new:
                    created += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠️  Order: {str(e)}'))

        if created:
            self.stdout.write(f'  ✅ {created} orders imported')

    def _import_json_statuslogs(self, logs):
        """Import status logs from JSON data."""
        created = 0
        for item in logs:
            try:
                order = Order.objects.filter(id=item.get('order')).first()
                if not order:
                    continue
                
                log, is_new = StatusLog.objects.get_or_create(
                    id=item.get('id'),
                    defaults={
                        'order': order,
                        'status_type': item.get('status_type'),
                        'old_status': item.get('old_status'),
                        'new_status': item.get('new_status'),
                    }
                )
                if is_new:
                    created += 1
            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  ⚠️  StatusLog: {str(e)}'))
        
        if created:
            self.stdout.write(f'  ✅ {created} status logs imported')

    def _import_json_order_items(self, items):
        """Import order items from JSON data."""
        to_create = []
        for item in items:
            order = Order.objects.filter(
                ticket_number=item.get('order_ticket')
            ).first()
            if not order:
                continue
            to_create.append(OrderItem(
                order=order,
                person_label=item.get('person_label', ''),
                meat=item.get('meat', ''),
                side=item.get('side', ''),
                vegetable=item.get('vegetable', ''),
                supplement=item.get('supplement', ''),
                supplement_price=item.get('supplement_price', 0),
                sort_order=item.get('sort_order', 0),
            ))
        if to_create:
            OrderItem.objects.bulk_create(to_create, ignore_conflicts=True)
            self.stdout.write(f'  ✅ {len(to_create)} order items imported')

    def _parse_date(self, value):
        """Parse date string to date object."""
        if not value:
            return None
        if isinstance(value, str):
            return datetime.fromisoformat(value).date()
        return value

    def _parse_datetime(self, value):
        """Parse datetime string to datetime object."""
        if not value:
            return None
        if isinstance(value, str):
            return datetime.fromisoformat(value)
        return value
