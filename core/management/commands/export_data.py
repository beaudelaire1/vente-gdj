"""
Export database data to Excel and JSON formats for backup/migration.

Usage:
    python manage.py export_data --format=excel --output=backup.xlsx
    python manage.py export_data --format=json --output=backup.json
    python manage.py export_data  # Default: current timestamp excel file
"""

import json
from datetime import datetime
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Model
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from core.models import Event, Customer, Order, OrderItem, StatusLog, UserProfile


class Command(BaseCommand):
    help = 'Export database data to Excel or JSON format'

    def add_arguments(self, parser):
        parser.add_argument(
            '--format',
            type=str,
            default='excel',
            choices=['excel', 'json'],
            help='Export format: excel or json (default: excel)',
        )
        parser.add_argument(
            '--output',
            type=str,
            default='',
            help='Output filename (default: data_YYYYMMDD_HHMMSS.ext)',
        )

    def handle(self, *args, **options):
        fmt = options['format']
        
        if fmt == 'excel':
            self.export_excel(options['output'])
        elif fmt == 'json':
            self.export_json(options['output'])

    def export_excel(self, filename=''):
        """Export to Excel with multiple sheets."""
        if not filename:
            now = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'data_{now}.xlsx'

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Export Events
        self.export_events_sheet(wb)
        
        # Export Customers
        self.export_customers_sheet(wb)
        
        # Export Orders
        self.export_orders_sheet(wb)
        
        # Export StatusLogs
        self.export_statuslogs_sheet(wb)

        # Export OrderItems
        self.export_orderitems_sheet(wb)

        try:
            wb.save(filename)
            self.stdout.write(
                self.style.SUCCESS(f'✅ Data exported to {filename}')
            )
        except Exception as e:
            raise CommandError(f'Failed to export to {filename}: {str(e)}')

    def export_events_sheet(self, wb):
        """Export Event data to worksheet."""
        ws = wb.create_sheet('Events')
        
        headers = ['ID', 'Name', 'Date', 'Description', 'Is Active', 'Created At']
        ws.append(headers)
        self._style_header(ws)
        
        for event in Event.objects.all():
            ws.append([
                event.id,
                event.name,
                event.date.isoformat(),
                event.description,
                event.is_active,
                event.created_at.isoformat(),
            ])

    def export_customers_sheet(self, wb):
        """Export Customer data to worksheet."""
        ws = wb.create_sheet('Customers')
        
        headers = ['ID', 'Name', 'Phone', 'Email', 'Created At']
        ws.append(headers)
        self._style_header(ws)
        
        for customer in Customer.objects.all():
            ws.append([
                customer.id,
                customer.name,
                customer.phone,
                customer.email,
                customer.created_at.isoformat(),
            ])

    def export_orders_sheet(self, wb):
        """Export Order data to worksheet."""
        ws = wb.create_sheet('Orders')
        
        headers = [
            'ID', 'Event', 'Customer', 'Ticket Number', 'QR Token',
            'Forfait', 'Nb Persons', 'Dining Type', 'Meat', 'Side',
            'Vegetable', 'Unit Price', 'Total Amount', 'Payment Status',
            'Prep Status', 'Started At', 'Served At', 'Created At'
        ]
        ws.append(headers)
        self._style_header(ws)
        
        for order in Order.objects.select_related('event', 'customer'):
            ws.append([
                order.id,
                order.event.name if order.event else '',
                order.customer.name if order.customer else '',
                order.ticket_number,
                str(order.qr_token),
                order.forfait,
                order.nb_persons,
                order.dining_type,
                order.meat,
                order.side,
                order.vegetable,
                float(order.unit_price),
                float(order.total_amount),
                order.payment_status,
                order.preparation_status,
                order.started_at.isoformat() if order.started_at else '',
                order.served_at.isoformat() if order.served_at else '',
                order.created_at.isoformat(),
            ])

    def export_statuslogs_sheet(self, wb):
        """Export StatusLog data to worksheet."""
        ws = wb.create_sheet('StatusLogs')
        
        headers = ['ID', 'Order', 'Status Type', 'Old Status', 'New Status', 'Created At']
        ws.append(headers)
        self._style_header(ws)
        
        for log in StatusLog.objects.select_related('order'):
            ws.append([
                log.id,
                f"#{log.order.ticket_number}" if log.order else '',
                log.status_type,
                log.old_status,
                log.new_status,
                log.created_at.isoformat(),
            ])

    def export_orderitems_sheet(self, wb):
        """Export OrderItem data to worksheet."""
        ws = wb.create_sheet('OrderItems')

        headers = ['ID', 'Order Ticket', 'Person Label', 'Meat', 'Side',
                   'Vegetable', 'Supplement', 'Supplement Price', 'Sort Order']
        ws.append(headers)
        self._style_header(ws)

        for item in OrderItem.objects.select_related('order'):
            ws.append([
                item.id,
                item.order.ticket_number,
                item.person_label,
                item.meat,
                item.side,
                item.vegetable,
                item.supplement,
                float(item.supplement_price),
                item.sort_order,
            ])

    def _style_header(self, ws):
        """Style worksheet header row."""
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def export_json(self, filename=''):
        """Export to JSON format."""
        if not filename:
            now = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'data_{now}.json'

        data = {
            'events': self._serialize_queryset(Event.objects.all()),
            'customers': self._serialize_queryset(Customer.objects.all()),
            'orders': self._serialize_orders(),
            'order_items': self._serialize_order_items(),
            'statuslogs': self._serialize_queryset(StatusLog.objects.all()),
        }

        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            self.stdout.write(
                self.style.SUCCESS(f'✅ Data exported to {filename}')
            )
        except Exception as e:
            raise CommandError(f'Failed to export to {filename}: {str(e)}')

    def _serialize_orders(self):
        """Serialize orders with event/customer names for portable JSON."""
        result = []
        for order in Order.objects.select_related('event', 'customer'):
            result.append({
                'event': order.event.name if order.event else '',
                'customer': order.customer.name if order.customer else '',
                'ticket_number': order.ticket_number,
                'forfait': order.forfait,
                'nb_persons': order.nb_persons,
                'dining_type': order.dining_type,
                'meat': order.meat,
                'side': order.side,
                'vegetable': order.vegetable,
                'unit_price': float(order.unit_price),
                'total_amount': float(order.total_amount),
                'payment_status': order.payment_status,
                'preparation_status': order.preparation_status,
                'started_at': order.started_at.isoformat() if order.started_at else None,
                'served_at': order.served_at.isoformat() if order.served_at else None,
            })
        return result

    def _serialize_order_items(self):
        """Serialize order items with ticket reference."""
        result = []
        for item in OrderItem.objects.select_related('order'):
            result.append({
                'order_ticket': item.order.ticket_number,
                'person_label': item.person_label,
                'meat': item.meat,
                'side': item.side,
                'vegetable': item.vegetable,
                'supplement': item.supplement,
                'supplement_price': float(item.supplement_price),
                'sort_order': item.sort_order,
            })
        return result

    def _serialize_queryset(self, qs):
        """Convert queryset to JSON-serializable list of dicts."""
        result = []
        for obj in qs:
            data = {}
            for field in obj._meta.fields:
                value = getattr(obj, field.name)
                if hasattr(value, 'isoformat'):  # datetime/date
                    value = value.isoformat()
                elif isinstance(value, bool):
                    pass  # Keep as is
                elif not isinstance(value, (str, int, float, type(None))):
                    value = str(value)
                data[field.name] = value
            result.append(data)
        return result
