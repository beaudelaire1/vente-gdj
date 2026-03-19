import openpyxl
from datetime import date as dt_date
from decimal import Decimal
from django.core.management.base import BaseCommand
from core.models import Event, Customer, MenuOption, Order


class Command(BaseCommand):
    help = "Importe les commandes depuis un fichier Excel (.xlsx)"

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help="Chemin vers le fichier Excel")
        parser.add_argument(
            '--event-name', type=str, default='Restaurant Éphémère GDJ EEBC',
            help="Nom de l'événement",
        )
        parser.add_argument(
            '--event-date', type=str, default='2026-03-28',
            help="Date de l'événement (AAAA-MM-JJ)",
        )
        parser.add_argument(
            '--sheet', type=str, default='inscrits',
            help="Nom de la feuille Excel",
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help="Afficher sans importer",
        )

    def handle(self, *args, **options):
        filepath = options['file']
        sheet_name = options['sheet']
        dry_run = options['dry_run']

        wb = openpyxl.load_workbook(filepath, data_only=True)
        if sheet_name not in wb.sheetnames:
            self.stderr.write(f"Feuille '{sheet_name}' introuvable. Disponibles : {wb.sheetnames}")
            return

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

        # Créer l'événement
        event_date = dt_date.fromisoformat(options['event_date'])
        event, created = Event.objects.get_or_create(
            name=options['event_name'],
            defaults={'date': event_date, 'is_active': True},
        )
        if created:
            self.stdout.write(self.style.SUCCESS(f"Événement créé : {event}"))
        else:
            self.stdout.write(f"Événement existant : {event}")

        imported = 0
        skipped = 0

        for i, row in enumerate(rows, start=2):
            nom = row[0]
            if not nom:
                continue

            forfait_raw = str(row[1] or '').strip().lower()
            nb_persons = int(row[2] or 1)
            viande = str(row[3] or '').strip()
            accompagnement = str(row[4] or '').strip()
            legume = str(row[5] or '').strip() if row[5] else ''
            type_raw = str(row[6] or '').strip()
            paiement_raw = str(row[8] or '').strip()

            # Déterminer le ticket (colonne K = index 10)
            ticket_raw = row[10]
            if ticket_raw is None or (isinstance(ticket_raw, str) and not ticket_raw.strip()):
                ticket_number = i  # fallback to row number
            else:
                try:
                    ticket_number = int(ticket_raw)
                except (ValueError, TypeError):
                    ticket_number = i

            # Mapper forfait
            forfait = Order.FORFAIT_FAMILLE if 'famille' in forfait_raw else Order.FORFAIT_INDIVIDUEL

            # Mapper type
            if 'emporter' in type_raw.lower():
                dining_type = Order.DINING_A_EMPORTER
            else:
                dining_type = Order.DINING_SUR_PLACE

            # Mapper paiement
            if 'payé' in paiement_raw.lower() and 'non' not in paiement_raw.lower():
                if 'attente' in paiement_raw.lower():
                    payment_status = Order.PAYMENT_ATTENTE
                else:
                    payment_status = Order.PAYMENT_PAYE
            else:
                payment_status = Order.PAYMENT_NON_PAYE

            if dry_run:
                self.stdout.write(
                    f"  #{ticket_number} {nom} | {viande} + {accompagnement} "
                    f"| {nb_persons}p | {forfait} | {dining_type} | {payment_status}"
                )
                imported += 1
                continue

            # Vérifier si déjà importé
            if Order.objects.filter(event=event, ticket_number=ticket_number).exists():
                skipped += 1
                continue

            # Créer le client
            customer, _ = Customer.objects.get_or_create(
                name=nom.strip(),
            )

            MenuOption.objects.get_or_create(
                option_type=MenuOption.TYPE_MEAT,
                label=(viande or 'Non spécifié').strip(),
            )
            MenuOption.objects.get_or_create(
                option_type=MenuOption.TYPE_SIDE,
                label=(accompagnement or 'Non spécifié').strip(),
            )
            if legume:
                MenuOption.objects.get_or_create(
                    option_type=MenuOption.TYPE_VEGETABLE,
                    label=legume.strip(),
                )

            # Créer la commande
            order = Order(
                event=event,
                customer=customer,
                ticket_number=ticket_number,
                forfait=forfait,
                nb_persons=nb_persons,
                dining_type=dining_type,
                meat=viande or 'Non spécifié',
                side=accompagnement or 'Non spécifié',
                vegetable=legume,
                unit_price=Decimal('0'),
                total_amount=Decimal('0'),
                payment_status=payment_status,
                preparation_status=Order.PREP_NON_LANCE,
            )
            order.compute_total()
            order.save()
            imported += 1

        action = "Prévu" if dry_run else "Importé"
        self.stdout.write(self.style.SUCCESS(
            f"\n{action} : {imported} commandes | Ignoré : {skipped}"
        ))
